#!/usr/bin/env python3
"""
Excel generate script for payroll results.
Creates a spreadsheet with cross-sheet references.
"""

import json
import sys

try:
      from openpyxl import Workbook
      from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
      from openpyxl.utils import get_column_letter
except ImportError:
      print("Error: openpyxl is not installed.")
      print("Install with: pip install openpyxl")
      sys.exit(1)


def create_master_sheet(wb, employees, grade_table):
      """Create master data sheet"""
      ws = wb.create_sheet("Master")

    # Header
      headers = ["ID", "Name", "Department", "Grade", "Base Salary", "Commute", "Dependents"]
      for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)

      # Employee data
      for row, emp in enumerate(employees, 2):
                ws.cell(row=row, column=1, value=emp['employee_id'])
                ws.cell(row=row, column=2, value=emp['employee_name'])
                ws.cell(row=row, column=3, value=emp['department'])
                ws.cell(row=row, column=4, value=emp['grade'])
                ws.cell(row=row, column=5, value=emp['base_salary'])
                ws.cell(row=row, column=6, value=emp['commute_allowance'])
                ws.cell(row=row, column=7, value=emp['dependents'])

      # Grade table
      ws.cell(row=6, column=1, value="Grade Table")
      ws.cell(row=7, column=1, value="Grade")
      ws.cell(row=7, column=2, value="Insurance Rate")
      ws.cell(row=7, column=3, value="Base Deduction")

    row = 8
    for grade, info in grade_table.items():
              ws.cell(row=row, column=1, value=grade)
              ws.cell(row=row, column=2, value=info['insurance_rate'])
              ws.cell(row=row, column=3, value=info['base_deduction'])
              row += 1

    return ws


def create_attendance_sheet(wb, employees):
      """Create attendance sheet"""
      ws = wb.create_sheet("Attendance")

    headers = ["ID", "Regular OT", "Late Night OT", "Holiday", "Holiday Night", "Absence", "Tardiness"]
    for col, header in enumerate(headers, 1):
              cell = ws.cell(row=1, column=col, value=header)
              cell.font = Font(bold=True)

    for row, emp in enumerate(employees, 2):
              att = emp['attendance']
              ws.cell(row=row, column=1, value=emp['employee_id'])
              ws.cell(row=row, column=2, value=att['regular_overtime_hours'])
              ws.cell(row=row, column=3, value=att['late_night_overtime_hours'])
              ws.cell(row=row, column=4, value=att['holiday_work_hours'])
              ws.cell(row=row, column=5, value=att['holiday_late_night_hours'])
              ws.cell(row=row, column=6, value=att['absence_days'])
              ws.cell(row=row, column=7, value=att['tardiness_count'])

    return ws


def create_allowance_sheet(wb, employees):
      """Create allowance calculation sheet"""
      ws = wb.create_sheet("Allowances")

    headers = ["ID", "Hourly Rate", "Regular OT", "Late Night", "Holiday", "Holiday Night", "Total"]
    for col, header in enumerate(headers, 1):
              cell = ws.cell(row=1, column=col, value=header)
              cell.font = Font(bold=True)

    for row, emp in enumerate(employees, 2):
              allow = emp['allowances']
              ws.cell(row=row, column=1, value=emp['employee_id'])
              ws.cell(row=row, column=2, value=emp['hourly_rate'])
              ws.cell(row=row, column=3, value=allow['regular_overtime'])
              ws.cell(row=row, column=4, value=allow['late_night'])
              ws.cell(row=row, column=5, value=allow['holiday_work'])
              ws.cell(row=row, column=6, value=allow['holiday_late_night'])
              ws.cell(row=row, column=7, value=allow['total'])

    return ws


def create_deduction_sheet(wb, employees):
      """Create deduction calculation sheet"""
      ws = wb.create_sheet("Deductions")

    headers = ["ID", "Absence", "Tardiness", "Total Deduct", "Social Ins", "Income Tax", "Statutory Total"]
    for col, header in enumerate(headers, 1):
              cell = ws.cell(row=1, column=col, value=header)
              cell.font = Font(bold=True)

    for row, emp in enumerate(employees, 2):
              ded = emp['deductions_from_pay']
              stat = emp['statutory_deductions']
              ws.cell(row=row, column=1, value=emp['employee_id'])
              ws.cell(row=row, column=2, value=ded['absence'])
              ws.cell(row=row, column=3, value=ded['tardiness'])
              ws.cell(row=row, column=4, value=ded['total'])
              ws.cell(row=row, column=5, value=stat['social_insurance'])
              ws.cell(row=row, column=6, value=stat['income_tax'])
              ws.cell(row=row, column=7, value=stat['total'])

    return ws


def create_payslip_sheet(wb, employees):
      """Create payslip summary sheet"""
      ws = wb.create_sheet("Payslip")

    headers = ["ID", "Name", "Base Salary", "Allowances", "Deductions", "Gross Pay", "Statutory", "Net Pay"]
    for col, header in enumerate(headers, 1):
              cell = ws.cell(row=1, column=col, value=header)
              cell.font = Font(bold=True)

    for row, emp in enumerate(employees, 2):
              ws.cell(row=row, column=1, value=emp['employee_id'])
              ws.cell(row=row, column=2, value=emp['employee_name'])
              ws.cell(row=row, column=3, value=emp['base_salary'])
              ws.cell(row=row, column=4, value=emp['allowances']['total'])
              ws.cell(row=row, column=5, value=emp['deductions_from_pay']['total'])
              ws.cell(row=row, column=6, value=emp['gross_pay'])
              ws.cell(row=row, column=7, value=emp['statutory_deductions']['total'])
              ws.cell(row=row, column=8, value=emp['net_pay'])

    # Totals
    total_row = len(employees) + 2
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{total_row-1})")
    ws.cell(row=total_row, column=7, value=f"=SUM(G2:G{total_row-1})")
    ws.cell(row=total_row, column=8, value=f"=SUM(H2:H{total_row-1})")

    return ws


def create_verification_sheet(wb, employees):
      """Create verification sheet"""
      ws = wb.create_sheet("Verification")

    headers = ["ID", "Name", "Expected Net", "Calculated Net", "Difference", "Status"]
    for col, header in enumerate(headers, 1):
              cell = ws.cell(row=1, column=col, value=header)
              cell.font = Font(bold=True)

    for row, emp in enumerate(employees, 2):
              ws.cell(row=row, column=1, value=emp['employee_id'])
              ws.cell(row=row, column=2, value=emp['employee_name'])
              ws.cell(row=row, column=3, value=emp['net_pay'])
              ws.cell(row=row, column=4, value=f"=Payslip!H{row}")
              ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
              ws.cell(row=row, column=6, value=f'=IF(E{row}=0,"OK","ERROR")')

    return ws


def generate_excel(input_file, output_file):
      """Generate Excel file from JSON data"""
      with open(input_file, 'r', encoding='utf-8') as f:
                data = json.load(f)

      employees = data['results']
      grade_table = data['grade_table']

    wb = Workbook()
    wb.remove(wb.active)

    create_master_sheet(wb, employees, grade_table)#!/usr/bin/env python3
"""
Excel generate script for payroll results.
Creates a spreadsheet with cross-sheet references.
"""

import json
import sys

try:
      from openpyxl import Workbook
      from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
      from openpyxl.utils import get_column_letter
except ImportError:
      print("Error: openpyxl is not installed.")
      print("Install with: pip install openpyxl")
      sys.exit(1)


def create_master_sheet(wb, employees, grade_table):
      """Create master data sheet"""
      ws = wb.create_sheet("Master")

    # Header
      headers = ["ID", "Name", "Department", "Grade", "Base Salary", "Commute", "Dependents"]
      for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)

      # Employee data
      for row, emp in enumerate(employees, 2):
                ws.cell(row=row, column=1, value=emp['employee_id'])
                ws.cell(row=row, column=2, value=emp['employee_name'])
                ws.cell(row=row, column=3, value=emp['department'])
                ws.cell(row=row, column=4, value=emp['grade'])
                ws.cell(row=row, column=5, value=emp['base_salary'])
                ws.cell(row=row, column=6, value=emp['commute_allowance'])
                ws.cell(row=row, column=7, value=emp['dependents'])

    # Grade table
    ws.cell(row=6, column=1, value="Grade Table")
    ws.cell(row=7, column=1, value="Grade")
    ws.cell(row=7, column=2, value="Insurance Rate")
    ws.cell(row=7, column=3, value="Base Deduction")

    row = 8
    for grade, info in grade_table.items():
              ws.cell(row=row, column=1, value=grade)
              ws.cell(row=row, column=2, value=info['insurance_rate'])
              ws.cell(row=row, column=3, value=info['base_deduction'])
              row += 1

    return ws


def create_attendance_sheet(wb, employees):
      """Create attendance sheet"""
      ws = wb.create_sheet("Attendance")

    headers = ["ID", "Regular OT", "Late Night OT", "Holiday", "Holiday Night", "Absence", "Tardiness"]
    for col, header in enumerate(headers, 1):
              cell = ws.cell(row=1, column=col, value=header)
              cell.font = Font(bold=True)

    for row, emp in enumerate(employees, 2):
              att = emp['attendance']
              ws.cell(row=row, column=1, value=emp['employee_id'])
              ws.cell(row=row, column=2, value=att['regular_overtime_hours'])
              ws.cell(row=row, column=3, value=att['late_night_overtime_hours'])
              ws.cell(row=row, column=4, value=att['holiday_work_hours'])
              ws.cell(row=row, column=5, value=att['holiday_late_night_hours'])
              ws.cell(row=row, column=6, value=att['absence_days'])
              ws.cell(row=row, column=7, value=att['tardiness_count'])

    return ws


def create_allowance_sheet(wb, employees):
      """Create allowance calculation sheet"""
      ws = wb.create_sheet("Allowances")

    headers = ["ID", "Hourly Rate", "Regular OT", "Late Night", "Holiday", "Holiday Night", "Total"]
    for col, header in enumerate(headers, 1):
              cell = ws.cell(row=1, column=col, value=header)
              cell.font = Font(bold=True)

    for row, emp in enumerate(employees, 2):
              allow = emp['allowances']
              ws.cell(row=row, column=1, value=emp['employee_id'])
              ws.cell(row=row, column=2, value=emp['hourly_rate'])
              ws.cell(row=row, column=3, value=allow['regular_overtime'])
              ws.cell(row=row, column=4, value=allow['late_night'])
              ws.cell(row=row, column=5, value=allow['holiday_work'])
              ws.cell(row=row, column=6, value=allow['holiday_late_night'])
              ws.cell(row=row, column=7, value=allow['total'])

    return ws


def create_deduction_sheet(wb, employees):
      """Create deduction calculation sheet"""
      ws = wb.create_sheet("Deductions")

    headers = ["ID", "Absence", "Tardiness", "Total Deduct", "Social Ins", "Income Tax", "Statutory Total"]
    for col, header in enumerate(headers, 1):
              cell = ws.cell(row=1, column=col, value=header)
              cell.font = Font(bold=True)

    for row, emp in enumerate(employees, 2):
              ded = emp['deductions_from_pay']
              stat = emp['statutory_deductions']
              ws.cell(row=row, column=1, value=emp['employee_id'])
              ws.cell(row=row, column=2, value=ded['absence'])
              ws.cell(row=row, column=3, value=ded['tardiness'])
              ws.cell(row=row, column=4, value=ded['total'])
              ws.cell(row=row, column=5, value=stat['social_insurance'])
              ws.cell(row=row, column=6,
